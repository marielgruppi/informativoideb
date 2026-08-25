"""
Processa os arquivos oficiais do Inep (divulgação Ideb por município e por escola,
2025) cruzando com o crosswalk RGInt/SRE, e gera uma planilha de indicadores
agregados prontos para os gráficos do informativo Ideb-MG.

Fontes (baixadas do Inep e do IBGE, na raiz do repo):
  - divulgacao_{anos_iniciais,anos_finais,ensino_medio}_municipios_2025.xlsx
  - divulgacao_{anos_iniciais,anos_finais,ensino_medio}_escolas_2025.xlsx
  - RGInt_SRE_MUN_UF.xlsx (aba DTB_Municípios_SRE_MG: cruzamento município -> RGInt/SRE)

Saída: analise/indicadores_ideb_mg.xlsx
"""
import re
import openpyxl
import pandas as pd

UF_VALIDAS = {
    "AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG","PA","PB",
    "PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO",
}

ETAPAS = {
    "anos_iniciais": "Anos Iniciais do Ensino Fundamental",
    "anos_finais": "Anos Finais do Ensino Fundamental",
    "ensino_medio": "Ensino Médio",
}


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("-", "", "ND*", "ND**", "ND***", "ND"):
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def load_divulgacao(path, nivel):
    """nivel: 'municipio' ou 'escola'. Retorna DataFrame long: uma linha por
    (chave, rede, ano) com colunas P, N, IDEB, META."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[9]

    col = {c: i for i, c in enumerate(codes) if c}
    idx_uf = col["SG_UF"]
    idx_mun_cod = col["CO_MUNICIPIO"]
    idx_mun_nome = col["NO_MUNICIPIO"]
    idx_rede = col["REDE"]
    if nivel == "escola":
        idx_esc_cod = col["ID_ESCOLA"]
        idx_esc_nome = col["NO_ESCOLA"]

    # mapeia ano -> índices de P, N, IDEB
    anos = {}
    for code, i in col.items():
        m = re.match(r"VL_(OBSERVADO|NOTA_MEDIA|INDICADOR_REND|PROJECAO)_(\d+)", code)
        if not m:
            continue
        kind, ano = m.groups()
        ano = int(re.sub(r"\D", "", ano)[:4])  # "20212" -> "2021"
        anos.setdefault(ano, {})[kind] = i

    records = []
    for row in rows[10:]:
        uf = row[idx_uf]
        if uf not in UF_VALIDAS:
            continue
        base = {
            "UF": uf,
            "CO_MUNICIPIO": row[idx_mun_cod],
            "NO_MUNICIPIO": row[idx_mun_nome],
            "REDE": row[idx_rede],
        }
        if nivel == "escola":
            base["CO_ESCOLA"] = row[idx_esc_cod]
            base["NO_ESCOLA"] = row[idx_esc_nome]
        for ano, idxs in anos.items():
            ideb = to_num(row[idxs["OBSERVADO"]]) if "OBSERVADO" in idxs else None
            n = to_num(row[idxs["NOTA_MEDIA"]]) if "NOTA_MEDIA" in idxs else None
            p = to_num(row[idxs["INDICADOR_REND"]]) if "INDICADOR_REND" in idxs else None
            meta = to_num(row[idxs["PROJECAO"]]) if "PROJECAO" in idxs else None
            if ideb is None and n is None and p is None:
                continue
            rec = dict(base)
            rec.update(ANO=ano, IDEB=ideb, N=n, P=p, META=meta)
            records.append(rec)
    df = pd.DataFrame.from_records(records)
    return df


REGIOES = {"Norte", "Nordeste", "Sudeste", "Sul", "Centro-Oeste"}

NOME_PARA_UF = {
    "Rondônia": "RO", "Acre": "AC", "Amazonas": "AM", "Roraima": "RR", "Pará": "PA",
    "Amapá": "AP", "Tocantins": "TO", "Maranhão": "MA", "Piauí": "PI", "Ceará": "CE",
    "R. G. do Norte": "RN", "Paraíba": "PB", "Pernambuco": "PE", "Alagoas": "AL",
    "Sergipe": "SE", "Bahia": "BA", "Minas Gerais": "MG", "Espírito Santo": "ES",
    "Rio de Janeiro": "RJ", "São Paulo": "SP", "Paraná": "PR", "Santa Catarina": "SC",
    "R. G. do Sul": "RS", "M. G. do Sul": "MS", "Mato Grosso": "MT", "Goiás": "GO",
    "Distrito Federal": "DF",
}


def normaliza_rede(v):
    if v is None:
        return None
    return re.sub(r"\s*\(\d+\)", "", str(v)).strip()


def load_regiao_uf_com_codigos(path, sheet):
    """Para arquivos com a linha de códigos VL_* (ex.: divulgacao_regioes_ufs_ideb_2025.xlsx)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[9]
    col = {c: i for i, c in enumerate(codes) if c}
    idx_nome = 0
    idx_rede = 1

    anos = {}
    for code, i in col.items():
        m = re.match(r"VL_(OBSERVADO|NOTA_MEDIA|INDICADOR_REND|PROJECAO)_(\d+)", code)
        if not m:
            continue
        kind, ano = m.groups()
        ano = int(re.sub(r"\D", "", ano)[:4])
        anos.setdefault(ano, {})[kind] = i

    records = []
    for row in rows[10:]:
        nome = row[idx_nome]
        if nome not in NOME_PARA_UF and nome not in REGIOES:
            continue
        base = {
            "NOME": nome,
            "UF": NOME_PARA_UF.get(nome),
            "REGIAO": nome if nome in REGIOES else None,
            "REDE": normaliza_rede(row[idx_rede]),
        }
        for ano, idxs in anos.items():
            ideb = to_num(row[idxs["OBSERVADO"]]) if "OBSERVADO" in idxs else None
            n = to_num(row[idxs["NOTA_MEDIA"]]) if "NOTA_MEDIA" in idxs else None
            p = to_num(row[idxs["INDICADOR_REND"]]) if "INDICADOR_REND" in idxs else None
            if ideb is None and n is None and p is None:
                continue
            rec = dict(base)
            rec.update(ANO=ano, IDEB=ideb, N=n, P=p)
            records.append(rec)
    return pd.DataFrame.from_records(records)


def load_regiao_uf_legado(path, sheet):
    """Para arquivos antigos sem linha de códigos (ex.: divulgacao_regioes_ufs_ideb_2019.xlsx),
    identificando as colunas pelo forward-fill do cabeçalho (linhas 6/7/8)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    r6, r7, r8 = rows[6], rows[7], rows[8]
    ncol = len(r6)

    # forward-fill do rótulo do bloco (linha 6)
    bloco = [None] * ncol
    atual = None
    for i in range(ncol):
        if r6[i] is not None:
            atual = r6[i]
        bloco[i] = atual

    idx_ideb, idx_n, idx_p = {}, {}, {}
    for i in range(ncol):
        b = bloco[i]
        if not b:
            continue
        m = re.match(r"IDEB\n(\d+)", str(b))
        if m and r6[i] is not None:  # só a própria coluna de início do bloco IDEB
            idx_ideb[int(m.group(1))] = i
            continue
        m = re.match(r"Taxa de Aprovação - (\d+)", str(b))
        if m and r8[i] == "Indicador de Rendimento (P)":
            idx_p[int(m.group(1))] = i
            continue
        m = re.match(r"Nota SAEB - (\d+)", str(b))
        if m and r7[i] == "Nota Média Padronizada (N)":
            idx_n[int(m.group(1))] = i

    anos = sorted(set(idx_ideb) | set(idx_n) | set(idx_p))
    records = []
    for row in rows[10:]:
        nome = row[0]
        if nome not in NOME_PARA_UF and nome not in REGIOES:
            continue
        base = {
            "NOME": nome,
            "UF": NOME_PARA_UF.get(nome),
            "REGIAO": nome if nome in REGIOES else None,
            "REDE": normaliza_rede(row[1]),
        }
        for ano in anos:
            ideb = to_num(row[idx_ideb[ano]]) if ano in idx_ideb else None
            n = to_num(row[idx_n[ano]]) if ano in idx_n else None
            p = to_num(row[idx_p[ano]]) if ano in idx_p else None
            if ideb is None and n is None and p is None:
                continue
            rec = dict(base)
            rec.update(ANO=ano, IDEB=ideb, N=n, P=p)
            records.append(rec)
    return pd.DataFrame.from_records(records)


def load_brasil_com_codigos(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[9]
    col = {c: i for i, c in enumerate(codes) if c}
    idx_rede = col["rede"]

    anos = {}
    for code, i in col.items():
        m = re.match(r"VL_(OBSERVADO|NOTA_MEDIA|INDICADOR_REND|PROJECAO)_(\d+)", code)
        if not m:
            continue
        kind, ano = m.groups()
        ano = int(re.sub(r"\D", "", ano)[:4])
        anos.setdefault(ano, {})[kind] = i

    records = []
    for row in rows[10:]:
        if row[0] != "Brasil":
            continue
        base = {"NOME": "Brasil", "UF": None, "REGIAO": None, "REDE": normaliza_rede(row[idx_rede])}
        for ano, idxs in anos.items():
            ideb = to_num(row[idxs["OBSERVADO"]]) if "OBSERVADO" in idxs else None
            n = to_num(row[idxs["NOTA_MEDIA"]]) if "NOTA_MEDIA" in idxs else None
            p = to_num(row[idxs["INDICADOR_REND"]]) if "INDICADOR_REND" in idxs else None
            if ideb is None and n is None and p is None:
                continue
            rec = dict(base)
            rec.update(ANO=ano, IDEB=ideb, N=n, P=p)
            records.append(rec)
    return pd.DataFrame.from_records(records)


SHEET_BRASIL = {
    "anos_iniciais": "Brasil (Anos Iniciais)",
    "anos_finais": "Brasil (Anos Finais)",
    "ensino_medio": "Brasil (EM)",
}

SHEET_UF = {
    "anos_iniciais": "UF e Regiões (AI)",
    "anos_finais": "UF e Regiões (AF)",
    "ensino_medio": "UF e Regiões (EM)",
}


def faixa_ideb(v):
    if v is None or pd.isna(v):
        return None
    if v < 4:
        return "< 4"
    if v < 5:
        return "4 a 4,9"
    if v < 6:
        return "5 a 5,9"
    return ">= 6"


def main():
    print("Carregando crosswalk RGInt/SRE...")
    cross = pd.read_excel("RGInt_SRE_MUN_UF.xlsx", sheet_name="DTB_Municípios_SRE_MG")
    cross = cross.rename(columns={
        "id_municipio": "CO_MUNICIPIO",
        "nome_municipio": "NO_MUNICIPIO_CROSS",
        "nome_rgint": "RGINT",
        "sre": "SRE",
    })[["CO_MUNICIPIO", "RGINT", "SRE"]]
    cross["CO_MUNICIPIO"] = cross["CO_MUNICIPIO"].astype(int)
    cross["SRE"] = cross["SRE"].str.replace("SRE ", "", regex=False).str.title()

    out = {}

    # ---- Série histórica por município (MG) — Gráficos 1 a 5 ----
    serie_frames = []
    for etapa_key, etapa_nome in ETAPAS.items():
        print(f"Lendo divulgacao_{etapa_key}_municipios_2025.xlsx ...")
        df = load_divulgacao(f"divulgacao_{etapa_key}_municipios_2025.xlsx", "municipio")
        df = df[df["UF"] == "MG"].copy()
        df["ETAPA"] = etapa_nome
        serie_frames.append(df)
    serie_mun = pd.concat(serie_frames, ignore_index=True)
    serie_mun = serie_mun.merge(cross, on="CO_MUNICIPIO", how="left")
    out["Serie_Municipios_MG"] = serie_mun

    # Média simples entre municípios por rede/etapa/ano — aproximação do "MG total"
    # (Inep calcula o agregado estadual a partir dos microdados pool, não como média
    # simples dos municípios; usar com essa ressalva no texto).
    approx_mg = (
        serie_mun.groupby(["ETAPA", "REDE", "ANO"])[["IDEB", "N", "P"]]
        .mean()
        .round(2)
        .reset_index()
        .sort_values(["ETAPA", "REDE", "ANO"])
    )
    out["MG_aprox_media_municipios"] = approx_mg

    # Variação P e N entre 2023 e 2025, por rede/etapa (Gráfico 5) — a partir da
    # mesma aproximação por média simples de municípios.
    piv = approx_mg[approx_mg["ANO"].isin([2023, 2025])].pivot_table(
        index=["ETAPA", "REDE"], columns="ANO", values=["N", "P"]
    )
    piv.columns = [f"{v}_{a}" for v, a in piv.columns]
    piv = piv.reset_index()
    piv["VAR_N_23_25"] = (piv["N_2025"] - piv["N_2023"]).round(3)
    piv["VAR_P_23_25"] = (piv["P_2025"] - piv["P_2023"]).round(3)
    out["Variacao_N_P_2023_2025"] = piv

    # ---- Nível escola: % de escolas por faixa de Ideb, por SRE/RGInt (2025) ----
    faixa_frames = []
    for etapa_key, etapa_nome in ETAPAS.items():
        print(f"Lendo divulgacao_{etapa_key}_escolas_2025.xlsx ...")
        df = load_divulgacao(f"divulgacao_{etapa_key}_escolas_2025.xlsx", "escola")
        df = df[(df["UF"] == "MG") & (df["ANO"] == 2025)].copy()
        df["ETAPA"] = etapa_nome
        faixa_frames.append(df)
    escolas_2025 = pd.concat(faixa_frames, ignore_index=True)
    escolas_2025 = escolas_2025.merge(cross, on="CO_MUNICIPIO", how="left")
    escolas_2025["FAIXA"] = escolas_2025["IDEB"].apply(faixa_ideb)
    out["Escolas_2025_MG"] = escolas_2025

    def tabela_faixas(df, rede, dimensao):
        sub = df[(df["REDE"] == rede) & df["FAIXA"].notna()].copy()
        tab = (
            sub.groupby(["ETAPA", dimensao, "FAIXA"])
            .size()
            .rename("N_ESCOLAS")
            .reset_index()
        )
        totais = tab.groupby(["ETAPA", dimensao])["N_ESCOLAS"].transform("sum")
        tab["PCT_ESCOLAS"] = (tab["N_ESCOLAS"] / totais * 100).round(1)
        return tab.sort_values(["ETAPA", dimensao, "FAIXA"])

    out["Faixas_Estadual_por_SRE"] = tabela_faixas(escolas_2025, "Estadual", "SRE")
    out["Faixas_Estadual_por_RGInt"] = tabela_faixas(escolas_2025, "Estadual", "RGINT")
    out["Faixas_Municipal_por_SRE"] = tabela_faixas(
        escolas_2025[escolas_2025["ETAPA"] != "Ensino Médio"], "Municipal", "SRE"
    )
    out["Faixas_Municipal_por_RGInt"] = tabela_faixas(
        escolas_2025[escolas_2025["ETAPA"] != "Ensino Médio"], "Municipal", "RGINT"
    )

    # N de escolas por SRE/rede/etapa, para aplicar corte de N mínimo depois se quiserem
    contagem = (
        escolas_2025.groupby(["ETAPA", "REDE", "SRE"])
        .size()
        .rename("N_ESCOLAS_TOTAL")
        .reset_index()
    )
    out["Contagem_Escolas_por_SRE"] = contagem

    # ---- Nível UF/Região — Gráfico 1, série completa de EM, e ranking nacional ----
    uf_frames = []
    for etapa_key, etapa_nome in ETAPAS.items():
        print(f"Lendo divulgacao_regioes_ufs_ideb_2025.xlsx [{SHEET_UF[etapa_key]}] ...")
        df = load_regiao_uf_com_codigos("divulgacao_regioes_ufs_ideb_2025.xlsx", SHEET_UF[etapa_key])
        df["ETAPA"] = etapa_nome
        uf_frames.append(df)
    serie_uf = pd.concat(uf_frames, ignore_index=True)

    # Estende a série do Ensino Médio para 2005-2015 usando o arquivo legado de 2019
    # (o Inep só calcula Ideb de Ensino Médio por município/escola a partir de 2017;
    # em nível de UF/região a série antiga volta a 2005).
    print("Lendo divulgacao_regioes_ufs_ideb_2019.xlsx [EM, anos 2005-2015] ...")
    em_legado = load_regiao_uf_legado("divulgacao_regioes_ufs_ideb_2019.xlsx", "UF e Regiões (EM)")
    em_legado = em_legado[em_legado["ANO"] <= 2015].copy()
    em_legado["ETAPA"] = "Ensino Médio"
    serie_uf = pd.concat([serie_uf, em_legado], ignore_index=True)

    brasil_frames = []
    for etapa_key, etapa_nome in ETAPAS.items():
        print(f"Lendo divulgacao_brasil_ideb_2025.xlsx [{SHEET_BRASIL[etapa_key]}] ...")
        df = load_brasil_com_codigos("divulgacao_brasil_ideb_2025.xlsx", SHEET_BRASIL[etapa_key])
        df["ETAPA"] = etapa_nome
        brasil_frames.append(df)
    serie_uf = pd.concat([serie_uf] + brasil_frames, ignore_index=True)

    serie_uf = serie_uf.sort_values(["ETAPA", "REGIAO", "UF", "REDE", "ANO"])
    out["Serie_UF_Regiao"] = serie_uf

    # Série só de Minas Gerais em nível UF (substitui a aproximação por média de
    # municípios como fonte confiável para citar "o Ideb de Minas Gerais foi X").
    serie_mg_uf = serie_uf[serie_uf["UF"] == "MG"].copy()
    out["Serie_MG_nivel_UF"] = serie_mg_uf

    # Ranking nacional das 27 UFs, por etapa/rede/ano — posição de MG.
    def ranking(df, rede, ano, etapa):
        sub = df[(df["REDE"] == rede) & (df["ANO"] == ano) & (df["ETAPA"] == etapa) & df["UF"].notna()]
        sub = sub.dropna(subset=["IDEB"]).sort_values("IDEB", ascending=False).reset_index(drop=True)
        sub["POSICAO"] = sub.index + 1
        return sub[["POSICAO", "UF", "IDEB"]]

    ranking_rows = []
    for etapa_nome in ETAPAS.values():
        for ano in (2023, 2025):
            rk = ranking(serie_uf, "Estadual", ano, etapa_nome)
            rk["ETAPA"] = etapa_nome
            rk["ANO"] = ano
            ranking_rows.append(rk)
    ranking_df = pd.concat(ranking_rows, ignore_index=True)[["ETAPA", "ANO", "POSICAO", "UF", "IDEB"]]
    out["Ranking_Estadual_por_UF"] = ranking_df

    mg_posicao = ranking_df[ranking_df["UF"] == "MG"].sort_values(["ETAPA", "ANO"])
    out["Ranking_MG_resumo"] = mg_posicao

    print("Salvando analise/indicadores_ideb_mg.xlsx ...")
    import os
    os.makedirs("analise", exist_ok=True)
    with pd.ExcelWriter("analise/indicadores_ideb_mg.xlsx", engine="openpyxl") as writer:
        for name, df in out.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    print("OK")


if __name__ == "__main__":
    main()
